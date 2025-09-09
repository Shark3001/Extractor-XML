import os
import io
import xml.etree.ElementTree as ET
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import openpyxl
from openpyxl.styles import PatternFill
import logging
from functools import wraps
import re

# Configuración de logging para una mejor depuración
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Configuración de la aplicación ---
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv("SECRET_KEY_APP_XML", "CAMBIA_ESTA_CLAVE_EN_RENDER")
app.config['CORRECT_PASSWORD'] = os.getenv("APP_PASSWORD", "AFC2024*")

# --- Funciones de utilidad ---

def formatear_numero(valor):
    """Formatea un valor numérico para su representación en Excel (punto a coma)."""
    if valor is None:
        return ""
    # Usa un regex para reemplazar el primer punto (si existe) y luego las comas
    # con puntos para evitar conflictos, luego la coma final.
    return str(valor).replace('.', ',', 1)

def formatear_fecha(fecha_str):
    """Formatea una cadena de fecha ISO 8601 a 'dd-mm-yyyy'."""
    if not fecha_str:
        return ""
    try:
        # Manejo de múltiples formatos ISO, incluyendo el 'Z'
        date_obj = datetime.fromisoformat(re.sub(r'Z$', '+00:00', fecha_str))
        return date_obj.strftime('%d-%m-%Y')
    except ValueError as e:
        logging.error(f"Error al formatear la fecha '{fecha_str}': {e}")
        return fecha_str

def login_required(f):
    """Decorador para proteger rutas que requieren autenticación."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            flash('Por favor, inicia sesión para acceder a esta función.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def parse_xml_value(element, xpath, default_value=""):
    """
    Función de utilidad para obtener el texto de un elemento XML de forma segura.
    Simplifica la lógica de manejo de 'None'.
    """
    elem = element.find(xpath)
    if elem is not None and elem.text is not None:
        return elem.text.strip()
    return default_value

def parse_and_format_xml_value(element, xpath, formatter=None, default_value=""):
    """
    Obtiene y formatea un valor de un elemento XML.
    Simplifica la lógica de manejo de 'None' y aplica un formateador opcional.
    """
    value = parse_xml_value(element, xpath, default_value)
    if formatter and value != default_value:
        return formatter(value)
    return value

def extract_invoice_data(root, filename):
    """
    Extrae los datos comunes y de resumen de un solo archivo XML.
    Retorna un diccionario con los datos extraídos.
    """
    if root.tag.split('}')[1] == "MensajeHacienda":
        logging.info(f"Saltando archivo {filename}: Es un MensajeHacienda.")
        return None, None

    # Normalizar el nombre de las etiquetas eliminando el namespace
    for elem in root.iter():
        elem.tag = elem.tag.split('}', 1)[-1]
    
    tipo_documento = root.tag
    
    # Extracción de datos comunes
    common_data = {
        "Clave": parse_xml_value(root, 'Clave'),
        "Consecutivo": parse_xml_value(root, 'NumeroConsecutivo'),
        "Fecha": formatear_fecha(parse_xml_value(root, 'FechaEmision')),
        "Nombre Emisor": parse_xml_value(root, 'Emisor/Nombre'),
        "Número Emisor": parse_xml_value(root, 'Emisor/Identificacion/Numero'),
        "Nombre Receptor": parse_xml_value(root, 'Receptor/Nombre'),
        "Número Receptor": parse_xml_value(root, 'Receptor/Identificacion/Numero'),
        "Archivo": filename,
        "Tipo de Documento": tipo_documento,
    }

    # Extracción de datos del resumen
    resumen_factura = root.find('ResumenFactura')
    resumen_data = {
        "Total Exento": parse_and_format_xml_value(resumen_factura, 'TotalExento', formatear_numero),
        "Total Exonerado": parse_and_format_xml_value(resumen_factura, 'TotalExonerado', formatear_numero),
        "Total Venta": parse_and_format_xml_value(resumen_factura, 'TotalVenta', formatear_numero),
        "Total Descuentos": parse_and_format_xml_value(resumen_factura, 'TotalDescuentos', formatear_numero),
        "Total Venta Neta": parse_and_format_xml_value(resumen_factura, 'TotalVentaNeta', formatear_numero),
        "Total Impuesto": parse_and_format_xml_value(resumen_factura, 'TotalImpuesto', formatear_numero),
        "Total Comprobante": parse_and_format_xml_value(resumen_factura, 'TotalComprobante', formatear_numero),
        "Otros Cargos": parse_and_format_xml_value(root.find('OtrosCargos'), 'MontoCargo', formatear_numero, "0,00"),
    }
    
    return common_data, resumen_data

def extract_detailed_line_data(root, common_data):
    """
    Extrae los datos de cada línea de detalle de un archivo XML.
    Retorna una lista de diccionarios.
    """
    detalles = []
    detalles_servicio = root.find('DetalleServicio')

    if detalles_servicio is None:
        return detalles

    resumen_factura = root.find('ResumenFactura')
    codigo_moneda = parse_xml_value(resumen_factura, 'CodigoTipoMoneda/CodigoMoneda')
    tipo_cambio = parse_and_format_xml_value(resumen_factura, 'CodigoTipoMoneda/TipoCambio', formatear_numero)
    total_gravado = parse_and_format_xml_value(resumen_factura, 'TotalGravado', formatear_numero)
    
    for linea in detalles_servicio.findall('LineaDetalle'):
        impuesto = linea.find('Impuesto')
        
        line_data = {
            "Código Cabys": parse_xml_value(linea, 'Codigo'),
            "Detalle": parse_xml_value(linea, 'Detalle'),
            "Cantidad": parse_and_format_xml_value(linea, 'Cantidad', formatear_numero),
            "Precio Unitario": parse_and_format_xml_value(linea, 'PrecioUnitario', formatear_numero),
            "Monto Total": parse_and_format_xml_value(linea, 'MontoTotal', formatear_numero),
            "Monto Descuento": parse_and_format_xml_value(linea, 'Descuento/MontoDescuento', formatear_numero, "0,00"),
            "Subtotal": parse_and_format_xml_value(linea, 'SubTotal', formatear_numero),
            "Tarifa (%)": parse_and_format_xml_value(impuesto, 'Tarifa', formatear_numero, "0,00"),
            "Monto Impuesto": parse_and_format_xml_value(impuesto, 'Monto', formatear_numero, "0,00"),
            "Impuesto Neto": parse_and_format_xml_value(linea, 'ImpuestoNeto', formatear_numero),
            "Código Moneda": codigo_moneda,
            "Tipo Cambio": tipo_cambio,
            "Total Gravado": total_gravado,
            # Se añaden datos del resumen a cada línea para la tabla detallada
            **{k: v for k, v in common_data.items() if k not in ["Archivo", "Tipo de Documento"]},
            "Total Exento": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalExento', formatear_numero),
            "Total Exonerado": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalExonerado', formatear_numero),
            "Total Venta": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalVenta', formatear_numero),
            "Total Descuentos": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalDescuentos', formatear_numero),
            "Total Venta Neta": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalVentaNeta', formatear_numero),
            "Total Impuesto": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalImpuesto', formatear_numero),
            "Total Comprobante": parse_and_format_xml_value(root.find('ResumenFactura'), 'TotalComprobante', formatear_numero),
            "Otros Cargos": parse_and_format_xml_value(root.find('OtrosCargos'), 'MontoCargo', formatear_numero, "0,00"),
            "Archivo": common_data["Archivo"],
            "Tipo de Documento": common_data["Tipo de Documento"]
        }
        
        detalles.append(line_data)
        
    return detalles

def extraer_datos_xml_en_memoria(xml_files, numero_receptor_filtro):
    """
    Procesa una lista de archivos XML y genera un archivo Excel en memoria.
    """
    wb = openpyxl.Workbook()

    # --- HOJA facturas_detalladas ---
    ws_detalladas = wb.active
    ws_detalladas.title = "facturas_detalladas"
    headers_detalladas = [
        "Clave", "Consecutivo", "Fecha", "Nombre Emisor", "Número Emisor", "Nombre Receptor", "Número Receptor",
        "Código Cabys", "Detalle", "Cantidad", "Precio Unitario", "Monto Total", "Monto Descuento", "Subtotal",
        "Tarifa (%)", "Monto Impuesto", "Impuesto Neto", "Código Moneda", "Tipo Cambio",
        "Total Gravado", "Total Exento", "Total Exonerado", "Total Venta", "Total Descuentos",
        "Total Venta Neta", "Total Impuesto", "Total Comprobante", "Otros Cargos", "Archivo", "Tipo de Documento"
    ]
    ws_detalladas.append(headers_detalladas)

    # --- HOJA facturas_resumidas ---
    ws_resumidas = wb.create_sheet(title="facturas_resumidas")
    headers_resumidas = [
        "Clave", "Consecutivo", "Fecha", "Nombre Emisor", "Número Emisor", "Número Receptor",
        "Total Exento", "Total Exonerado", "Total Venta", "Total Descuentos", "Total Venta Neta",
        "Total Impuesto", "Total Comprobante", "Otros Cargos", "Archivo", "Tipo de Documento"
    ]
    ws_resumidas.append(headers_resumidas)

    for uploaded_file in xml_files:
        filename = uploaded_file.filename
        try:
            tree = ET.parse(uploaded_file)
            root = tree.getroot()
            
            common_data, resumen_data = extract_invoice_data(root, filename)
            if common_data is None:
                continue

            # --- Escribir datos resumidos ---
            fila_resumida = [common_data[h] for h in headers_resumidas if h in common_data or h in resumen_data]
            ws_resumidas.append(fila_resumida)

            # --- Escribir datos detallados ---
            detalles = extract_detailed_line_data(root, common_data)
            for detalle in detalles:
                fila_detallada = [detalle[h] for h in headers_detalladas]
                ws_detalladas.append(fila_detallada)
                
        except ET.ParseError as e:
            flash(f"Error de formato XML en '{filename}': {e}", 'error')
            logging.error(f"Error de formato XML en '{filename}': {e}")
        except Exception as e:
            flash(f"Error inesperado al procesar '{filename}': {e}", 'error')
            logging.error(f"Error inesperado al procesar '{filename}': {e}")

    # --- Formato de celdas ---
    fill_celeste = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_rojo = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
    
    # Formato para facturas detalladas
    for col_name in ["Consecutivo", "Fecha", "Detalle", "Tarifa (%)", "Total Gravado", "Total Venta", "Total Venta Neta", "Total Comprobante"]:
        if col_name in headers_detalladas:
            col_idx = headers_detalladas.index(col_name) + 1
            for cell in list(ws_detalladas.columns)[col_idx-1]:
                cell.fill = fill_celeste
                
    # Formato para facturas resumidas
    for col_name in ["Consecutivo", "Fecha", "Total Impuesto", "Total Venta Neta", "Total Venta", "Total Exento", "Total Exonerado", "Total Comprobante"]:
        if col_name in headers_resumidas:
            col_idx = headers_resumidas.index(col_name) + 1
            for cell in list(ws_resumidas.columns)[col_idx-1]:
                cell.fill = fill_celeste

    # Colorear en rojo si el número de receptor no coincide
    if numero_receptor_filtro:
        receptor_col_detalladas = headers_detalladas.index("Número Receptor") + 1
        for row in ws_detalladas.iter_rows(min_row=2):
            if str(row[receptor_col_detalladas - 1].value).strip() != str(numero_receptor_filtro).strip():
                for cell in row:
                    cell.fill = fill_rojo
                    
        receptor_col_resumidas = headers_resumidas.index("Número Receptor") + 1
        for row in ws_resumidas.iter_rows(min_row=2):
            if str(row[receptor_col_resumidas - 1].value).strip() != str(numero_receptor_filtro).strip():
                for cell in row:
                    cell.fill = fill_rojo

    # Guardar en un buffer de memoria
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# --- Rutas de la aplicación Flask ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == app.config['CORRECT_PASSWORD']:
            session['logged_in'] = True
            flash('Inicio de sesión exitoso. ¡Bienvenido!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Contraseña incorrecta. Inténtalo de nuevo.', 'error')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/logout')
@login_required
def logout():
    session.pop('logged_in', None)
    flash('Has cerrado sesión correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
@login_required
def upload_files():
    if 'xml_files' not in request.files or not request.files.getlist('xml_files'):
        flash('No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('index'))

    files = request.files.getlist('xml_files')
    numero_receptor = request.form.get('numero_receptor')

    if not numero_receptor or not numero_receptor.strip():
        flash('El número de identificación del receptor es obligatorio.', 'error')
        return redirect(url_for('index'))

    excel_stream = extraer_datos_xml_en_memoria(files, numero_receptor)

    return send_file(
        excel_stream,
        download_name='datos_facturas.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    # Usar un puerto dinámico si está disponible (Render) o el 5000 por defecto
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
